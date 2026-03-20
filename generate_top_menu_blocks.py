#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import re
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


GEN_BLOCK_START = "<!-- generated_top_menu_block:start -->"
GEN_BLOCK_END = "<!-- generated_top_menu_block:end -->"

HTTP_RE = re.compile(r"^https?://", re.IGNORECASE)
WS_RE = re.compile(r"\s+")
DIV_TAG_RE = re.compile(r"<(/?)div\b[^>]*>", re.IGNORECASE)
FOOTERS_DIV_RE = re.compile(
    r'<div\b[^>]*class=["\'][^"\']*footers_static_data[^"\']*["\'][^>]*>',
    re.IGNORECASE,
)

TOP_MENU_EXISTING_STYLE_TAG = (
    '<style data-id="top-menu-existing-style">'
    '.footers_static_data .losb-body{height:auto!important;max-height:none!important;min-height:0!important;overflow:visible!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"]{position:relative!important;overflow:visible!important;padding-top:2px!important;width:100%!important;max-width:100%!important;display:flex!important;align-items:stretch!important;gap:14px!important;box-sizing:border-box!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-level-left{flex:0 0 24%!important;min-width:150px!important;max-width:260px!important;height:auto!important;max-height:none!important;overflow-y:visible!important;overflow-x:visible!important;display:flex!important;flex-direction:column!important;padding-right:12px!important;border-right:1px solid #a9aea8!important;box-sizing:border-box!important;float:none!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-element-sub{display:block!important;padding:2px 8px 2px 6px!important;text-decoration:none!important;color:#4e554d!important;line-height:1.3!important;border:1px solid transparent!important;font-size:16px!important;background:transparent!important;cursor:pointer!important;text-align:left!important;width:100%!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-element-sub > a{display:block!important;color:inherit!important;text-decoration:none!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-element-sub.active{border-color:#95bfc0!important;background:#f4fbfb!important;color:#5f8d8c!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-level-right{flex:1 1 auto!important;min-width:0!important;width:auto!important;max-width:none!important;height:auto!important;max-height:none!important;overflow:visible!important;padding-left:14px!important;box-sizing:border-box!important;position:relative!important;float:none!important;display:block!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-sub{width:100%!important;max-width:100%!important;min-width:0!important;height:auto!important;max-height:none!important;overflow:visible!important;position:absolute!important;left:100%!important;top:0!important;opacity:0!important;pointer-events:none!important;transition:left .18s ease-in-out,opacity .18s ease-in-out!important;display:block!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-sub.active{position:relative!important;left:0!important;opacity:1!important;pointer-events:auto!important;height:auto!important;max-height:none!important;overflow:visible!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-result{display:inline-block!important;width:33.333%!important;max-width:33.333%!important;vertical-align:top!important;box-sizing:border-box!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-result:not(:last-child){border-right:1px solid #a9aea8!important;padding-right:18px!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-result:not(:first-child){padding-left:18px!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-link{display:block!important;text-decoration:none!important;color:#5f665f!important;line-height:1.3!important;font-size:16px!important;white-space:normal!important;overflow:visible!important;text-overflow:clip!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] p{margin:0 0 4px 0!important;display:block!important;width:100%!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-link:hover{color:#2f342f!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"]::after{content:"";display:block;clear:both}'
    '@media (max-width:900px){'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-level-left{width:100%!important;max-width:100%!important;min-width:auto!important;border-right:none!important;padding-right:0!important;margin-bottom:10px!important;float:none!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-level-right{width:100%!important;max-width:100%!important;padding-left:0!important;float:none!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-sub{position:absolute!important;left:100%!important;top:0!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-content-sub.active{position:relative!important;left:0!important}'
    '.footers_static_data .losb-content[data-id="losb-content-top_menu"] .losb-menu-result{display:block;width:100%!important;border-right:none!important;padding:0 0 10px 0!important}'
    "}"
    "</style>"
)


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = WS_RE.sub(" ", text).strip()
    return text


def cut_domain(url_or_path: str) -> str:
    value = clean_text(url_or_path)
    if not value:
        return ""

    if HTTP_RE.match(value):
        parsed = urlparse(value)
        value = parsed.path or "/"

    if not value.startswith("/"):
        value = f"/{value}"

    value = re.sub(r"/+", "/", value)
    if value != "/":
        value = value.rstrip("/")
    return value or "/"


def normalize_uri(url_or_path: str) -> str:
    value = cut_domain(url_or_path)
    return "/" if value == "" else value


def split_columns(items: list[tuple[str, str]], column_count: int = 4) -> list[list[tuple[str, str]]]:
    if not items:
        return []

    column_count = max(1, min(column_count, len(items)))
    items_per_column = int(math.ceil(len(items) / column_count))

    columns: list[list[tuple[str, str]]] = []
    for idx in range(0, len(items), items_per_column):
        columns.append(items[idx : idx + items_per_column])
    return columns


def parse_section_parts(path_value: object) -> list[str]:
    text = clean_text(path_value)
    if not text:
        return []
    return [part for part in (clean_text(p) for p in text.split("/")) if part]


def build_hierarchy_context(products_path: Path, category_links: list[tuple[str, str]]) -> dict:
    name_to_url: dict[str, str] = {}
    url_to_name: dict[str, str] = {}
    name_order: dict[str, int] = {}
    fallback_names: list[str] = []

    for idx, (name, href) in enumerate(category_links):
        if name not in name_to_url:
            name_to_url[name] = href
        uri = normalize_uri(href)
        if uri not in url_to_name:
            url_to_name[uri] = name
        if name not in name_order:
            name_order[name] = idx
            fallback_names.append(name)

    edge_counts: dict[str, Counter] = defaultdict(Counter)
    wb = load_workbook(products_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [clean_text(v) for v in headers_row]
    section_col = 10
    for candidate in ["Раздел", "Розділ", "Разделы", "Розділи"]:
        if candidate in headers:
            section_col = headers.index(candidate) + 1
            break

    for (section_value,) in ws.iter_rows(
        min_row=2,
        max_row=ws.max_row,
        min_col=section_col,
        max_col=section_col,
        values_only=True,
    ):
        parts = parse_section_parts(section_value)
        if len(parts) < 2:
            continue

        for idx in range(len(parts) - 1):
            parent = parts[idx]
            child = parts[idx + 1]
            edge_counts[parent][child] += 1

    wb.close()

    children: dict[str, list[str]] = {}
    parent_best: dict[str, tuple[str, int]] = {}

    for parent, counter in edge_counts.items():
        sorted_children = sorted(
            counter.keys(),
            key=lambda name: (name_order.get(name, 10**9), -counter[name], name.lower()),
        )
        children[parent] = sorted_children

        for child, count in counter.items():
            prev = parent_best.get(child)
            if prev is None:
                parent_best[child] = (parent, count)
                continue

            prev_parent, prev_count = prev
            if count > prev_count or (
                count == prev_count
                and name_order.get(parent, 10**9) < name_order.get(prev_parent, 10**9)
            ):
                parent_best[child] = (parent, count)

    parent_of = {child: parent for child, (parent, _) in parent_best.items()}

    return {
        "name_to_url": name_to_url,
        "url_to_name": url_to_name,
        "children": children,
        "parent_of": parent_of,
        "fallback_names": fallback_names,
    }


def find_root_name(name: str, parent_of: dict[str, str]) -> str:
    current = name
    seen: set[str] = set()
    while current in parent_of and current not in seen:
        seen.add(current)
        current = parent_of[current]
    return current


def unique_names(names: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for name in names:
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(name)
    return result


def names_to_links(
    names: list[str],
    name_to_url: dict[str, str],
    current_uri: str,
    limit: int,
) -> list[tuple[str, str]]:
    result: list[tuple[str, str]] = []
    seen: set[str] = set()

    for name in names:
        href = name_to_url.get(name)
        if not href:
            continue
        if normalize_uri(href) == current_uri:
            continue

        key = name.lower()
        if key in seen:
            continue
        seen.add(key)

        result.append((name, href))
        if len(result) >= limit:
            break

    return result


def build_menu_data(
    context: dict,
    current_url: str,
    current_name: str = "",
    section_path: str = "",
    left_limit: int = 7,
    right_limit: int = 24,
) -> dict | None:
    name_to_url: dict[str, str] = context["name_to_url"]
    url_to_name: dict[str, str] = context["url_to_name"]
    children: dict[str, list[str]] = context["children"]
    parent_of: dict[str, str] = context["parent_of"]
    fallback_names: list[str] = context["fallback_names"]

    current_uri = normalize_uri(current_url)

    resolved_name = clean_text(current_name)
    if not resolved_name:
        resolved_name = url_to_name.get(current_uri, "")
    if not resolved_name and section_path:
        parts = parse_section_parts(section_path)
        if parts:
            resolved_name = parts[-1]

    root_name = find_root_name(resolved_name, parent_of) if resolved_name else ""

    left_candidates: list[str] = []
    if root_name and root_name in children:
        left_candidates = children[root_name]
    elif resolved_name and resolved_name in children:
        left_candidates = children[resolved_name]
    elif resolved_name and resolved_name in parent_of:
        left_candidates = children.get(parent_of[resolved_name], [])

    if not left_candidates:
        left_candidates = [name for name in fallback_names if children.get(name)]

    left_candidates = [name for name in unique_names(left_candidates) if name in name_to_url]
    if resolved_name and resolved_name in left_candidates:
        left_candidates = [resolved_name] + [name for name in left_candidates if name != resolved_name]

    submenus: list[list[tuple[str, str]]] = []
    safe_right_limit = max(6, min(right_limit, 40))
    left_items: list[tuple[str, str]] = []

    for left_name in left_candidates:
        submenu_names = unique_names(children.get(left_name, []))
        if not submenu_names:
            parent = parent_of.get(left_name)
            if parent:
                submenu_names = [name for name in children.get(parent, []) if name != left_name]
        if not submenu_names and root_name and root_name in children:
            submenu_names = [name for name in children[root_name] if name != left_name]
        if not submenu_names:
            continue

        submenu_links = names_to_links(
            names=submenu_names,
            name_to_url=name_to_url,
            current_uri=current_uri,
            limit=safe_right_limit,
        )
        if not submenu_links:
            continue

        left_items.append((left_name, name_to_url[left_name]))
        submenus.append(submenu_links)

        if len(left_items) >= left_limit:
            break

    if len(left_items) < 2:
        return None

    return {"left": left_items, "submenus": submenus}


def render_submenu_table(links: list[tuple[str, str]], column_count: int = 3) -> str:
    columns = split_columns(links, column_count=column_count) if links else [[] for _ in range(column_count)]

    # Keep a stable 3-column layout for visual parity with the target screenshot.
    while len(columns) < column_count:
        columns.append([])

    html: list[str] = []
    html.append('<table class="losb-submenu-table"><tbody><tr>')
    for column in columns:
        html.append("<td><ul>")
        for anchor, href in column:
            html.append(f'<li><a class="losb-link" href="{href}">{anchor}</a></li>')
        html.append("</ul></td>")
    html.append("</tr></tbody></table>")
    return "".join(html)


def render_top_menu_block(items: dict | None) -> str:
    if not items:
        return ""

    left_items: list[tuple[str, str]] = items.get("left", [])
    submenu_groups: list[list[tuple[str, str]]] = items.get("submenus", [])
    if len(left_items) < 2:
        return ""

    left_count = len(left_items)
    while len(submenu_groups) < left_count:
        submenu_groups.append([])

    html: list[str] = []
    html.append(GEN_BLOCK_START)
    html.append('<div class="footers_static_data top-menu-static" style="margin-left:auto;margin-right:auto;max-width:1140px;">')
    html.append('<div class="losb-block">')
    html.append('<div class="losb-header"><div class="losb-menu">')
    html.append('<span class="losb-menu-element active" data-id="top_menu">Топ меню</span>')
    html.append("</div></div>")
    html.append('<div class="losb-body">')
    html.append('<div class="losb-content active" data-id="losb-content-top_menu">')
    html.append(f'<div class="losb-top-layout" data-left-count="{left_count}">')
    html.append('<div class="losb-content-level-left">')
    for idx, (anchor, _href) in enumerate(left_items):
        active_cls = " active" if idx == 0 else ""
        html.append(
            f'<button type="button" class="losb-menu-element-sub{active_cls}" data-index="{idx}">{anchor}</button>'
        )
    html.append("</div>")

    html.append('<div class="losb-content-level-right">')
    for idx, submenu in enumerate(submenu_groups):
        active_cls = " active" if idx == 0 else ""
        html.append(f'<div class="losb-content-sub{active_cls}" data-index="{idx}">')
        html.append(render_submenu_table(submenu, column_count=3))
        html.append("</div>")
    html.append("</div>")
    html.append("</div>")
    html.append("</div></div></div>")

    html.append(
        "<style>"
        ".top-menu-static .losb-block{display:flex;flex-direction:column;margin-top:10px;font-family:inherit;color:#586059}"
        ".top-menu-static .losb-header{margin-bottom:8px}"
        ".top-menu-static .losb-menu{display:flex;gap:8px;flex-wrap:wrap;align-items:center}"
        ".top-menu-static .losb-menu-element{display:inline-block;padding:7px 24px;border:1px solid #8f948d;background:#fff;color:#515950;font-size:16px;line-height:1.1}"
        ".top-menu-static .losb-menu-element.active{border-color:#95bfc0;background:#eef7f7;color:#6e9291}"
        ".top-menu-static .losb-body{margin-top:2px;overflow:hidden}"
        ".top-menu-static .losb-top-layout{display:flex;align-items:stretch}"
        ".top-menu-static .losb-content-level-left{width:24%;min-width:230px;display:flex;flex-direction:column;padding-right:12px;border-right:1px solid #a9aea8}"
        ".top-menu-static .losb-menu-element-sub{display:block;padding:2px 8px 2px 6px;text-decoration:none;color:#4e554d;line-height:1.3;border:1px solid transparent;font-size:16px;background:transparent;cursor:pointer;text-align:left;width:100%}"
        ".top-menu-static .losb-menu-element-sub.active{border-color:#95bfc0;background:#f4fbfb;color:#5f8d8c}"
        ".top-menu-static .losb-content-level-right{width:76%;padding-left:14px;position:relative;overflow:hidden}"
        ".top-menu-static .losb-content-sub{width:100%;position:absolute;left:100%;top:0;opacity:0;pointer-events:none;transition:left .18s ease-in-out,opacity .18s ease-in-out}"
        ".top-menu-static .losb-content-sub.active{position:relative;left:0;opacity:1;pointer-events:auto}"
        ".top-menu-static .losb-submenu-table{width:100%;table-layout:fixed;border-collapse:collapse}"
        ".top-menu-static .losb-submenu-table td{width:33.333%;vertical-align:top}"
        ".top-menu-static .losb-submenu-table td:not(:last-child){border-right:1px solid #a9aea8;padding-right:18px}"
        ".top-menu-static .losb-submenu-table td:not(:first-child){padding-left:18px}"
        ".top-menu-static .losb-submenu-table ul{margin:0;padding:0;list-style:none}"
        ".top-menu-static .losb-submenu-table li{margin:0;padding:0 0 4px 0}"
        ".top-menu-static .losb-link{display:block;text-decoration:none;color:#5f665f;line-height:1.3;font-size:16px}"
        ".top-menu-static .losb-link:hover{color:#2f342f}"
        "@media (max-width:900px){"
        ".top-menu-static .losb-top-layout{display:block}"
        ".top-menu-static .losb-content-level-left{width:100%;min-width:auto;border-right:none;padding-right:0;margin-bottom:10px}"
        ".top-menu-static .losb-content-level-right{width:100%;padding-left:0;position:relative}"
        ".top-menu-static .losb-content-sub{position:absolute;left:100%;top:0}"
        ".top-menu-static .losb-content-sub.active{position:relative;left:0}"
        ".top-menu-static .losb-submenu-table,.top-menu-static .losb-submenu-table tbody,.top-menu-static .losb-submenu-table tr{display:block}"
        ".top-menu-static .losb-submenu-table td{display:block;width:100%!important;border-right:none!important;padding:0 0 10px 0!important}"
        "}"
        "</style>"
    )
    html.append(
        "<script>"
        "(function(){"
        "const root=document.currentScript.closest('.top-menu-static');"
        "if(!root||root.dataset.topMenuInit==='1'){return;}"
        "root.dataset.topMenuInit='1';"
        "const menuItems=[...root.querySelectorAll('.losb-menu-element-sub')];"
        "const panes=[...root.querySelectorAll('.losb-content-sub')];"
        "if(!menuItems.length||!panes.length){return;}"
        "const setActive=(idx)=>{"
        "menuItems.forEach((el,i)=>el.classList.toggle('active',i===idx));"
        "panes.forEach((el)=>el.classList.toggle('active',parseInt(el.dataset.index||'-1',10)===idx));"
        "};"
        "menuItems.forEach((el)=>{el.addEventListener('click',(event)=>{event.preventDefault();const idx=parseInt(el.dataset.index||'0',10);setActive(isNaN(idx)?0:idx);});});"
        "setActive(0);"
        "})();"
        "</script>"
    )
    html.append("</div>")
    html.append(GEN_BLOCK_END)
    return "".join(html)


def find_matching_div_end(text: str, opening_div_pos: int) -> int | None:
    opening_match = DIV_TAG_RE.search(text, opening_div_pos)
    if not opening_match or opening_match.start() != opening_div_pos:
        return None

    depth = 1
    for div_match in DIV_TAG_RE.finditer(text, opening_match.end()):
        is_close = div_match.group(1) == "/"
        if is_close:
            depth -= 1
        else:
            depth += 1

        if depth == 0:
            return div_match.end()

    return None


def find_first_footers_static_block_range(text: str) -> tuple[int, int] | None:
    match = FOOTERS_DIV_RE.search(text)
    if not match:
        return None

    end_pos = find_matching_div_end(text, match.start())
    if end_pos is None:
        return None
    return (match.start(), end_pos)


def remove_div_by_data_id(text: str, data_id: str) -> str:
    pattern = re.compile(
        rf'<div\b[^>]*data-id=["\']{re.escape(data_id)}["\'][^>]*>',
        re.IGNORECASE,
    )
    match = pattern.search(text)
    if not match:
        return text

    end_pos = find_matching_div_end(text, match.start())
    if end_pos is None:
        return text

    return text[: match.start()] + text[end_pos:]


def render_top_menu_content_for_existing(menu_data: dict | None) -> str:
    if not menu_data:
        return ""

    left_items: list[tuple[str, str]] = menu_data.get("left", [])
    submenu_groups: list[list[tuple[str, str]]] = menu_data.get("submenus", [])
    if len(left_items) < 2:
        return ""

    while len(submenu_groups) < len(left_items):
        submenu_groups.append([])

    html: list[str] = []
    html.append('<div class="losb-content " data-id="losb-content-top_menu">')
    html.append('<div class="losb-content-level-left">')
    for idx, (name, href) in enumerate(left_items, start=1):
        active_cls = " active" if idx == 1 else ""
        if href:
            html.append(
                f'<div class="losb-menu-element-sub{active_cls}" data-id="{idx}"><a href="{href}">{name}</a></div>'
            )
        else:
            html.append(f'<div class="losb-menu-element-sub{active_cls}" data-id="{idx}">{name}</div>')
    html.append("</div>")
    html.append('<div class="losb-content-level-right">')

    for idx, links in enumerate(submenu_groups, start=1):
        active_cls = " active" if idx == 1 else ""
        columns = split_columns(links, column_count=3) if links else [[], [], []]
        while len(columns) < 3:
            columns.append([])

        html.append(f'<div class="losb-content-sub{active_cls}" data-id="losb-content-sub-{idx}">')
        for column in columns:
            html.append('<div class="losb-menu-result">')
            for anchor, href in column:
                html.append(f'<p><a class="losb-link" href="{href}">{anchor}</a></p>')
            html.append("</div>")
        html.append("</div>")

    html.append("</div>")
    html.append("</div>")
    return "".join(html)


def add_top_menu_to_existing_footer(block_html: str, menu_data: dict | None) -> str:
    top_menu_content = render_top_menu_content_for_existing(menu_data)
    if not top_menu_content:
        return block_html

    block = re.sub(
        r'<style\b[^>]*data-id=["\']top-menu-existing-style["\'][^>]*>.*?</style>',
        "",
        block_html,
        flags=re.IGNORECASE | re.DOTALL,
    )

    block = re.sub(
        r'<span\b[^>]*class=["\'][^"\']*losb-menu-element[^"\']*["\'][^>]*data-id=["\']top_menu["\'][^>]*>.*?</span>',
        "",
        block,
        flags=re.IGNORECASE | re.DOTALL,
    )
    block = remove_div_by_data_id(block, "losb-content-top_menu")

    menu_match = re.search(
        r'<div\b[^>]*class=["\'][^"\']*losb-menu[^"\']*["\'][^>]*>',
        block,
        flags=re.IGNORECASE,
    )
    if not menu_match:
        return block

    menu_close = block.find("</div>", menu_match.end())
    if menu_close == -1:
        return block
    menu_tab = '<span class="losb-menu-element " data-id="top_menu">Топ меню</span>'
    block = block[:menu_close] + menu_tab + block[menu_close:]

    # Keep styles outside .losb-block because project CSS has ".losb-block * { display:flex; }".
    # If style tag is inside .losb-block, its CSS text can become visible in the layout.
    footers_match = FOOTERS_DIV_RE.search(block)
    if footers_match:
        block = block[: footers_match.end()] + TOP_MENU_EXISTING_STYLE_TAG + block[footers_match.end() :]

    body_match = re.search(
        r'<div\b[^>]*class=["\'][^"\']*losb-body[^"\']*["\'][^>]*>',
        block,
        flags=re.IGNORECASE,
    )
    if not body_match:
        return block

    body_end = find_matching_div_end(block, body_match.start())
    if body_end is None:
        return block

    body_close_start = body_end - len("</div>")
    block = block[:body_close_start] + top_menu_content + block[body_close_start:]
    return block


def remove_generated_marker_block(text: str) -> str:
    if GEN_BLOCK_START not in text:
        return text

    pattern = re.compile(
        re.escape(GEN_BLOCK_START) + r".*?" + re.escape(GEN_BLOCK_END),
        re.DOTALL,
    )
    return pattern.sub("", text)


def remove_first_footers_static_block(text: str) -> str:
    block_range = find_first_footers_static_block_range(text)
    if not block_range:
        return text

    start, end = block_range
    return text[:start] + text[end:]


def merge_description(existing: object, block_html: str, menu_data: dict | None) -> str:
    text = "" if existing is None else str(existing)
    text = remove_generated_marker_block(text)

    block_range = find_first_footers_static_block_range(text)
    if block_range:
        start, end = block_range
        current_block = text[start:end]
        updated_block = add_top_menu_to_existing_footer(current_block, menu_data)
        return text[:start] + updated_block + text[end:]

    text = text.rstrip()
    return f"{text}{block_html}" if text and block_html else (block_html or text)


def read_category_links(path: Path) -> list[tuple[str, str]]:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    links: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row:
            continue

        anchor = clean_text(row[0] if len(row) >= 1 else "")
        url = clean_text(row[1] if len(row) >= 2 else "")
        if not anchor or not url:
            continue

        href = cut_domain(url)
        key = (anchor.lower(), normalize_uri(href))
        if key in seen:
            continue

        seen.add(key)
        links.append((anchor, href))

    wb.close()
    return links


def find_column_index(headers: list[str], candidates: list[str]) -> int:
    header_map = {h: idx for idx, h in enumerate(headers, start=1)}
    for candidate in candidates:
        if candidate in header_map:
            return header_map[candidate]
    raise KeyError(f"Column not found. Tried: {candidates}")


def process_products(
    input_path: Path,
    output_path: Path,
    context: dict,
    links_per_block: int,
) -> int:
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    header_values = [clean_text(cell.value) for cell in ws[1]]
    desc_col = find_column_index(header_values, ["Описание товара(UA)", "Описание товара(RU)"])
    alias_col = find_column_index(header_values, ["Алиас", "Alias"])
    try:
        section_col = find_column_index(header_values, ["Раздел", "Розділ"])
    except KeyError:
        section_col = 10

    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        alias = clean_text(ws.cell(row=row_idx, column=alias_col).value)
        current_uri = f"/{alias.strip('/')}" if alias else "/"
        section_path = clean_text(ws.cell(row=row_idx, column=section_col).value)

        menu_data = build_menu_data(
            context=context,
            current_url=current_uri,
            section_path=section_path,
            left_limit=7,
            right_limit=links_per_block,
        )
        block_html = render_top_menu_block(menu_data)
        if not block_html:
            continue

        old_description = ws.cell(row=row_idx, column=desc_col).value
        ws.cell(row=row_idx, column=desc_col).value = merge_description(old_description, block_html, menu_data)
        updated += 1

    wb.save(output_path)
    wb.close()
    return updated


def category_has_header(ws) -> bool:
    first_name = clean_text(ws.cell(row=1, column=1).value).lower()
    first_url = clean_text(ws.cell(row=1, column=2).value)
    if HTTP_RE.match(first_url):
        return False
    return first_name in {"name", "название", "назва", "category", "категория", "категорія"}


def process_categories(
    input_path: Path,
    output_path: Path,
    context: dict,
    links_per_block: int,
) -> int:
    wb = load_workbook(input_path)
    ws = wb[wb.sheetnames[0]]

    has_header = category_has_header(ws)
    start_row = 2 if has_header else 1

    desc_col = ws.max_column + 1
    if has_header:
        ws.cell(row=1, column=desc_col).value = "Описание товара(UA)"

    updated = 0
    for row_idx in range(start_row, ws.max_row + 1):
        name = clean_text(ws.cell(row=row_idx, column=1).value)
        url = clean_text(ws.cell(row=row_idx, column=2).value)
        if not name or not url:
            continue

        menu_data = build_menu_data(
            context=context,
            current_url=url,
            current_name=name,
            left_limit=7,
            right_limit=links_per_block,
        )
        block_html = render_top_menu_block(menu_data)
        if not block_html:
            continue

        old_description = ws.cell(row=row_idx, column=desc_col).value
        ws.cell(row=row_idx, column=desc_col).value = merge_description(old_description, block_html, menu_data)
        updated += 1

    wb.save(output_path)
    wb.close()
    return updated


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate static TOP menu HTML blocks into product/category XLSX descriptions."
    )
    parser.add_argument("--products", required=True, type=Path, help="Path to product XLSX file")
    parser.add_argument("--categories", required=True, type=Path, help="Path to category XLSX file")
    parser.add_argument(
        "--links-per-block",
        type=int,
        default=24,
        help="How many links to place in each right submenu block (default: 24)",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=Path("."),
        help="Directory for output files (default: current directory)",
    )
    args = parser.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)

    category_links = read_category_links(args.categories)
    if not category_links:
        raise RuntimeError("No category links were found in category XLSX.")
    context = build_hierarchy_context(args.products, category_links)

    products_out = args.out_dir / f"{args.products.stem}_with_top_menu.xlsx"
    categories_out = args.out_dir / f"{args.categories.stem}_with_top_menu.xlsx"

    product_rows = process_products(
        input_path=args.products,
        output_path=products_out,
        context=context,
        links_per_block=args.links_per_block,
    )
    category_rows = process_categories(
        input_path=args.categories,
        output_path=categories_out,
        context=context,
        links_per_block=args.links_per_block,
    )

    print(f"Done. Products updated: {product_rows}")
    print(f"Done. Categories updated: {category_rows}")
    print(f"Saved: {products_out}")
    print(f"Saved: {categories_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
